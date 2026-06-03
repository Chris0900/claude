import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "2026 글로벌 마케팅 캘린더"

# ── 색상 정의 ──
C = {
    'hdr':      '1F3864', 'hdr_f':    'FFFFFF',
    'vn_h':     '375623', 'vn_b':     'E2EFDA', 'vn_f':  '1E3A1E',
    'jp_h':     '843C0C', 'jp_b':     'FCE4D6', 'jp_f':  '5C2700',
    'us_h':     '1F4E79', 'us_b':     'DEEAF1', 'us_f':  '1F4E79',
    'cn_h':     '7F6000', 'cn_b':     'FFF2CC', 'cn_f':  '7F6000',
    'tw_h':     '375623', 'tw_b':     'DDEBF7', 'tw_f':  '1F4E79',
    'mn_h':     '595959', 'mn_b':     'F2F2F2', 'mn_f':  '595959',
    'sg_h':     '595959', 'sg_b':     'EAF4FB', 'sg_f':  '1A5276',
    'me_h':     '7B3F00', 'me_b':     'FEF9E7', 'me_f':  '7B3F00',
    'au_h':     '1C5E1C', 'au_b':     'F0FFF0', 'au_f':  '1C5E1C',
    'cm_h':     '404040', 'cm_b':     'F5F5F5', 'cm_f':  '404040',
    'goal':     'D9D9D9',
    'sub':      'BDD7EE', 'sub_f':    '1F3864',
    'tot':      '1F3864', 'tot_f':    'FFFFFF',
    'event':    'FFF2CC',
    'zero':     'F9F9F9',
    'num':      'FFFFFF',
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color='000000', size=9):
    return Font(bold=bold, color=color, size=size, name='맑은 고딕')

def align(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

thin = Side(style='thin', color='CCCCCC')
med  = Side(style='medium', color='999999')
thk  = Side(style='medium', color='404040')

def border(l=thin, r=thin, t=thin, b=thin):
    return Border(left=l, right=r, top=t, bottom=b)

MONTHS = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월']
KEY_EVENTS = {
    1: '신년·뗏준비', 2: '뗏·춘절', 3: '라마단', 4: '이드',
    5: '어린이날·520', 6: '618', 7: '프라임데이', 8: '대만어버이날·오봉',
    9: '중추절', 10: '광군절예열', 11: '광군절·블프', 12: '쌍12·연말'
}

# 섹션별 데이터 [국가, 국가색, 목표, [(활동명, 비고, [m1..m12])]]
PLAN = [
    ('공통', 'cm', '글로벌 브랜드 인지도 구축 (웰라이프 공식 인스타그램 운영)', [
        (' 글로벌 인스타그램 공식계정 운영',  '콘텐츠 제작+관리 월고정',
         [10,10,10,10,10,10,10,10,10,10,10,10]),
    ]),
    ('베트남', 'vn', '1차목표: 오프라인 판매량 증대 + CGF 브랜드 전문성 구축 |  대표 3개월 1회 현지 직접 방문', [
        (' [하이하이] P-Max 연계 디지털 광고',     '8월 리뉴얼 런칭 집중 / 10월 뉴케어 집중 / 총 3억',
         [0,0,0,0,0,20,50,100,20,80,30,0]),
        (' CGF 전문성 (의사협회 논문 + KOL/KOC)', '6월 논문착수 → 7~10월 KOL/KOC 순차 배포',
         [0,0,0,0,0,20,15,15,10,10,0,0]),
        (' 오프라인 엔드매대 + 시음행사',          '월 1천만 / 윈마트 262개 확장 연동',
         [0,0,0,0,0,10,10,15,15,10,10,10]),
        (' 이너샵 VMD + 홍보물 + 온라인',         '8월 오픈 연동 / 매장 VMD+SNS 병행',
         [0,0,0,0,0,0,0,20,10,0,0,0]),
    ]),
    ('일본', 'jp', '최종목표: 코스트코 데모 행사 효율 향상 + 뉴케어 지속 운영 |  6/5 활동 시작', [
        (' 인플루언서 마케팅 (인스타+광고코드+현장스케치 영상)', '행사 알림+브랜드 노출 / 6월 집중 런칭',
         [0,0,0,0,0,25,20,20,15,15,10,5]),
        (' 뉴케어 일본 인스타 브랜드페이지 + 체험단',          '월 5백만 고정 운영 / 리뷰 콘텐츠 누적',
         [0,0,0,0,0,5,5,5,5,5,5,5]),
        (' 칼비 후루그라 코스트코 공동 샘플링 협업',            '후루그라 검색→웰라이프 유입 전략',
         [0,0,0,0,0,20,0,0,15,0,0,0]),
        (' 홈쇼핑/라이브커머스/온라인 오픈 SA·DA',             '오픈 시점 집중 광고 집행',
         [0,0,0,0,0,0,20,0,0,0,20,0]),
    ]),
    ('미국', 'us', '최종목표: 시장진입 초기 인지도·주의 확보 | 대상아메리카 협업 예정', [
        (' 브랜드 전용존 구축 + 현장 시음 진행', '매대 전용 섹션 세팅 + 시음 에이전시',
         [0,0,0,0,0,25,15,10,0,0,0,0]),
        ('️ 블랙프라이데이·Cyber Monday 이벤트', '10월 준비 / 11월 집중 집행',
         [0,0,0,0,0,0,0,0,0,10,40,10]),
        (' 헤이N 브랜드 아마존 테스트 광고',    '아마존 Sponsored·PPC 테스트 집행',
         [0,0,0,0,0,0,5,15,15,5,5,5]),
    ]),
    ('중국', 'cn', '최종목표: 역직구 채널 안정화 + GP Top5 진입 | 화룬강중 신규 800 파이프라인 런칭', [
        (' 618 쇼핑절 KOL 라이브방송 + 역직구 집중', '6/18 D-day 집중 집행',
         [0,0,0,0,0,60,0,0,0,0,0,0]),
        (' KOL 영상마케팅 + 전문가(의사) 추천 연계', '7~10월 KOL 주기 배포',
         [0,0,0,0,0,0,20,25,25,10,0,0]),
        (' 광군절 11.11 집중 캠페인',               '10월 예열 + 11월 D-day',
         [0,0,0,0,0,0,0,0,0,20,60,0]),
        (' 쌍12 (12/12) 연말 마감 캠페인',          '연간 목표 달성 마감',
         [0,0,0,0,0,0,0,0,0,0,0,30]),
        (' 화룬강중 (신규) 런칭 지원',              '800 파이프라인 / 런칭 예산 우선 배정 / 케이코스메몰 축소분 재배분',
         [0,0,0,0,0,20,10,10,0,0,0,0]),
    ]),
    ('대만', 'tw', '최종목표: 음료시장 성수기 집중 + Medfirst FSMP Top3 달성', [
        ('️ 여름 음료 성수기 집중 광고 (7~9월)',       '여름 건강음료 수요 피크 시즌',
         [0,0,0,0,0,0,20,25,15,0,0,0]),
        (' Medfirst 전문약국 O4O 마케팅 (LINE→약국)', '닥터 추천 마크 부착 후 집행',
         [0,0,0,0,0,0,0,0,0,10,0,10]),
    ]),
    ('몽골', 'mn', '최종목표: 현지 자체 브랜딩 지원 + 스포츠 마케팅 강화 | 베트남 골든 없음 → 인서트 영상 지원 전환', [
        (' 스포츠 이벤트 협찬 (마라톤 수분섭취 조끼 등)', '현지 파트너 주도 / 웰라이프 지원',
         [0,0,0,0,0,5,0,0,5,0,0,10]),
        (' 인서트 영상 제작 + 현지 배포 지원',           '베트남 골든 채널 종료 → 몽골 자체 SNS 인서트로 전환',
         [0,0,0,0,0,0,20,20,10,0,0,0]),
    ]),
    ('싱가포르', 'sg', '최종목표: 가디언 입점 매출 확대 + 박람회 신규 채널 개척', [
        (' 가디언 협업 확대 + 미팅 후 입점 지원',  '가디언 83개점 → 확장 공략',
         [0,0,0,0,0,15,0,0,0,15,0,0]),
        (' 현지 박람회 / 전시회 참가',            '신규 바이어 발굴',
         [0,0,0,0,0,0,0,0,10,0,0,0]),
    ]),
    ('중동', 'me', ' 시장 분위기 악화 → 광고 최소화 / FOC 재협상 병행', [
        (' Dermazone 최소 프로모 지원',  '현상 유지 / 추가 투자 없음',
         [0,0,0,0,0,5,0,0,5,0,0,10]),
    ]),
    ('호주', 'au', '최종목표: 천호케어 파트너십 유지 + 대상 호주 채널 안정화', [
        (' 천호케어 + 대상 호주 공동 마케팅 지원', '파트너 요청 시 지원',
         [0,0,0,0,0,5,0,0,10,0,0,5]),
    ]),
]

# ── 열 너비 설정 ──
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 44
ws.column_dimensions['C'].width = 28
for i, m in enumerate(MONTHS):
    col = get_column_letter(4 + i)
    ws.column_dimensions[col].width = 7
ws.column_dimensions[get_column_letter(16)].width = 9  # 합계

# ══════════════════════════════════════
# Row 1: 타이틀
# ══════════════════════════════════════
r = 1
ws.row_dimensions[r].height = 28
ws.merge_cells(f'A{r}:P{r}')
c = ws.cell(r, 1, '2026년 글로벌 마케팅 캘린더  |  총 예산 15억원  |  단위: 백만원')
c.fill = fill(C['hdr']); c.font = Font(bold=True, color='FFFFFF', size=14, name='맑은 고딕')
c.alignment = align('center')

# ──────────────────────────────────────
# Row 2: 주요 이벤트 라인
# ──────────────────────────────────────
r = 2
ws.row_dimensions[r].height = 22
ws.cell(r, 1, '').fill = fill(C['hdr'])
ws.cell(r, 2, ' 월별 주요 이벤트').fill = fill(C['hdr'])
ws.cell(r, 2).font = Font(bold=True, color='FFD700', size=9, name='맑은 고딕')
ws.cell(r, 2).alignment = align('left')
ws.cell(r, 3, '').fill = fill(C['hdr'])
for i, m in enumerate(MONTHS):
    c = ws.cell(r, 4+i, KEY_EVENTS.get(i+1,''))
    c.fill = fill('FFD700' if i+1 in [6,9,11] else 'FFF2CC' if i+1 in [2,5,8,12] else 'F2F2F2')
    c.font = Font(bold=(i+1 in [6,9,11]), color='7F4F00', size=7, name='맑은 고딕')
    c.alignment = align('center')
ws.cell(r, 16, '').fill = fill(C['hdr'])

# ──────────────────────────────────────
# Row 3: 컬럼 헤더
# ──────────────────────────────────────
r = 3
ws.row_dimensions[r].height = 20
headers = ['국가', '활동 내용', '비고 / 집행 방향'] + MONTHS + ['연간합계']
for ci, h in enumerate(headers, 1):
    c = ws.cell(r, ci, h)
    c.fill = fill(C['hdr']); c.font = font(True, 'FFFFFF', 9)
    c.alignment = align('center')
    c.border = border(thk, thk, thk, thk)

# ══════════════════════════════════════
# 데이터 출력
# ══════════════════════════════════════
row = 4
grand_totals = [0]*12

for (nation, nc, goal, acts) in PLAN:
    hk   = nc + '_h'; bk = nc + '_b'; fk = nc + '_f'
    hcol = C.get(hk, '595959'); bcol = C.get(bk, 'F5F5F5'); fcol = C.get(fk, '000000')

    # ── 섹션 목표 행 ──
    ws.row_dimensions[row].height = 28
    c = ws.cell(row, 1, nation)
    c.fill = fill(hcol); c.font = Font(bold=True, color='FFFFFF', size=9, name='맑은 고딕')
    c.alignment = align('center')
    c.border = border(thk, thk, thk, thk)

    ws.merge_cells(f'B{row}:P{row}')
    c = ws.cell(row, 2, f'▶ 목표: {goal}')
    c.fill = fill(C['goal']); c.font = Font(bold=True, color=hcol, size=9, name='맑은 고딕')
    c.alignment = align('left', wrap=True)
    c.border = border(thk, thk, thk, thk)
    row += 1

    # ── 활동 행 ──
    section_monthly = [0]*12
    for (act_name, note, monthly) in acts:
        ws.row_dimensions[row].height = 18
        ws.cell(row, 1, '').fill = fill(bcol)
        ws.cell(row, 1).border = border(thk, thin, thin, thin)

        c = ws.cell(row, 2, act_name)
        c.fill = fill(bcol); c.font = font(False, fcol, 9)
        c.alignment = align('left'); c.border = border(thin, thin, thin, thin)

        c = ws.cell(row, 3, note)
        c.fill = fill(bcol); c.font = font(False, '595959', 8)
        c.alignment = align('left'); c.border = border(thin, thin, thin, thin)

        row_total = sum(monthly)
        for i, val in enumerate(monthly):
            c = ws.cell(row, 4+i, val if val else '')
            c.fill = fill(bcol if val else C['zero'])
            c.font = font(val > 0, fcol if val else 'CCCCCC', 9)
            c.alignment = align('center')
            c.border = border(thin, thin, thin, thin)
            if val: section_monthly[i] += val

        c = ws.cell(row, 16, row_total if row_total else '')
        c.fill = fill(bcol); c.font = font(True, fcol, 9)
        c.alignment = align('center'); c.border = border(thin, thk, thin, thin)
        row += 1

    # ── 소계 행 ──
    ws.row_dimensions[row].height = 18
    c = ws.cell(row, 1, nation)
    c.fill = fill(hcol); c.font = Font(bold=True, color='FFFFFF', size=8, name='맑은 고딕')
    c.alignment = align('center'); c.border = border(thk, thk, thk, thk)

    c = ws.cell(row, 2, '소계')
    c.fill = fill(C['sub']); c.font = font(True, C['sub_f'], 9)
    c.alignment = align('center'); c.border = border(thk, thin, thk, thk)

    ws.cell(row, 3, '').fill = fill(C['sub'])
    ws.cell(row, 3).border = border(thin, thin, thk, thk)

    sub_total = sum(section_monthly)
    for i, val in enumerate(section_monthly):
        c = ws.cell(row, 4+i, val if val else '')
        c.fill = fill(C['sub']); c.font = font(True, C['sub_f'], 9)
        c.alignment = align('center'); c.border = border(thin, thin, thk, thk)
        grand_totals[i] += val

    c = ws.cell(row, 16, sub_total if sub_total else '')
    c.fill = fill(C['sub']); c.font = font(True, C['sub_f'], 9)
    c.alignment = align('center'); c.border = border(thin, thk, thk, thk)
    row += 1

# ══════════════════════════════════════
# 총합계 행
# ══════════════════════════════════════
ws.row_dimensions[row].height = 22
ws.merge_cells(f'A{row}:C{row}')
c = ws.cell(row, 1, '총 합계 (백만원)')
c.fill = fill(C['tot']); c.font = Font(bold=True, color='FFFFFF', size=10, name='맑은 고딕')
c.alignment = align('center'); c.border = border(thk,thk,thk,thk)

grand_total = sum(grand_totals)
for i, val in enumerate(grand_totals):
    c = ws.cell(row, 4+i, val if val else '')
    c.fill = fill(C['tot']); c.font = Font(bold=True, color='FFD700', size=10, name='맑은 고딕')
    c.alignment = align('center'); c.border = border(thin,thin,thk,thk)

c = ws.cell(row, 16, grand_total)
c.fill = fill(C['tot']); c.font = Font(bold=True, color='FFD700', size=12, name='맑은 고딕')
c.alignment = align('center'); c.border = border(thin,thk,thk,thk)

# ── 예산 활용률 안내 행 ──
row += 1
ws.row_dimensions[row].height = 16
ws.merge_cells(f'A{row}:P{row}')
note_txt = (f'  ※ 총 집행 예산 {grand_total:,}백만원  |  목표 15억원 대비 {grand_total/1500*100:.1f}%  '
            f'|  상반기(1~6월) 합계: {sum(grand_totals[:6]):,}  |  하반기(7~12월) 합계: {sum(grand_totals[6:]):,}  '
            f'|  주요 집중월: 6월(618) / 9월(중추절) / 11월(광군절·블프)')
c = ws.cell(row, 1, note_txt)
c.fill = fill('FFFACD'); c.font = Font(italic=True, color='595959', size=8, name='맑은 고딕')
c.alignment = align('left', wrap=False)

# ── 페이지 설정 ──
ws.freeze_panes = 'D4'
ws.sheet_view.showGridLines = True
ws.print_title_rows = '1:3'

out = '/home/user/claude/2026_글로벌마케팅캘린더.xlsx'
wb.save(out)
print(f'저장 완료: {out}')
print(f'총 예산: {grand_total}백만원')
print(f'월별 합계: {dict(zip([str(i+1)+"월" for i in range(12)], grand_totals))}')
